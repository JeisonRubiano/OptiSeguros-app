
import openpyxl
import sys

file_path = r'c:\Users\Jeison\Documents\Proyectos trabajo\Proyeto Pasantia\Detalle  negocios nuevos y recaudos.xlsx'

try:
    print(f"Opening file: {file_path}")
    # read_only=True is much faster and uses less memory for large files
    wb = openpyxl.load_workbook(file_path, read_only=True, keep_links=False)
    print("Sheets found:")
    for sheet in wb.sheetnames:
        print(f"- {sheet}")
        # Get approximate row count for the first sheet to verify
        if sheet == wb.sheetnames[0] or 'DETALLE' in sheet.upper():
            ws = wb[sheet]
            print(f"  (Inspecting '{sheet}' dimensions...)")
            print(f"  Max row estimate: {ws.max_row}")
except Exception as e:
    print(f"Error: {e}")
